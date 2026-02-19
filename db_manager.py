"""
Модуль работы с базами данных (Excel/CSV)
"""

import os
from typing import Optional, List, Dict, Any
from dataclasses import dataclass, field
from datetime import datetime


@dataclass
class DatabaseRecord:
    """Запись из базы данных"""
    row_index: int
    data: Dict[str, Any]
    
    def get(self, column: str, default: Any = None) -> Any:
        """Получить значение из колонки"""
        return self.data.get(column, default)


@dataclass
class DatabaseSearchResult:
    """Результат поиска в БД"""
    found: bool
    record: Optional[DatabaseRecord] = None
    all_matches: List[DatabaseRecord] = field(default_factory=list)
    error: str = ""


class DatabaseManager:
    """Менеджер баз данных"""
    
    def __init__(self):
        self.databases: Dict[str, 'DatabaseConnection'] = {}
        self.last_search_results: List[DatabaseRecord] = []
    
    def add_database(self, name: str, filepath: str) -> bool:
        """Добавить базу данных"""
        try:
            if not os.path.exists(filepath):
                return False
            
            conn = DatabaseConnection(filepath)
            conn.connect()
            self.databases[name] = conn
            return True
        except Exception as e:
            print(f"Ошибка подключения к БД {name}: {e}")
            return False
    
    def remove_database(self, name: str) -> bool:
        """Удалить базу данных"""
        if name in self.databases:
            self.databases[name].close()
            del self.databases[name]
            return True
        return False
    
    def get_database(self, name: str) -> Optional['DatabaseConnection']:
        """Получить подключение к БД"""
        return self.databases.get(name)
    
    def search_in_database(
        self, 
        db_name: str, 
        search_column: str, 
        search_value: str,
        return_columns: List[str] = None
    ) -> DatabaseSearchResult:
        """
        Поиск записи в базе данных
        
        Args:
            db_name: Имя базы данных
            search_column: Колонка для поиска
            search_value: Искомое значение
            return_columns: Колонки для возврата (None = все)
        
        Returns:
            DatabaseSearchResult с результатами
        """
        db = self.databases.get(db_name)
        if not db:
            return DatabaseSearchResult(
                found=False,
                error=f"База данных '{db_name}' не найдена"
            )
        
        try:
            # Поиск первого совпадения
            record = db.find_row(search_column, search_value)
            
            if not record:
                return DatabaseSearchResult(found=False)
            
            # Поиск всех совпадений
            all_records = db.find_all_rows(search_column, search_value)
            
            self.last_search_results = all_records
            
            return DatabaseSearchResult(
                found=True,
                record=record,
                all_matches=all_records
            )
            
        except Exception as e:
            return DatabaseSearchResult(
                found=False,
                error=str(e)
            )
    
    def get_column_values(
        self,
        db_name: str,
        column: str,
        filter_column: str = None,
        filter_value: str = None
    ) -> List[str]:
        """
        Получить все значения из колонки
        
        Args:
            db_name: Имя базы данных
            column: Колонка для получения значений
            filter_column: Колонка для фильтрации (опционально)
            filter_value: Значение фильтра (опционально)
        """
        db = self.databases.get(db_name)
        if not db:
            return []
        
        try:
            return db.get_column_values(column, filter_column, filter_value)
        except Exception:
            return []
    
    def close_all(self):
        """Закрыть все подключения"""
        for db in self.databases.values():
            db.close()
        self.databases.clear()


class DatabaseConnection:
    """Подключение к базе данных (Excel/CSV)"""
    
    def __init__(self, filepath: str):
        self.filepath = filepath
        self.is_connected = False
        self.data = []  # Список словарей
        self.columns = []  # Названия колонок
        self._workbook = None
        self._sheet = None
        self._excel_app = None
    
    def connect(self) -> bool:
        """Подключиться к базе данных"""
        try:
            ext = os.path.splitext(self.filepath)[1].lower()
            
            if ext in ['.xlsx', '.xls']:
                return self._connect_excel()
            elif ext == '.csv':
                return self._connect_csv()
            else:
                raise ValueError(f"Неподдерживаемый формат: {ext}")
                
        except Exception as e:
            print(f"Ошибка подключения: {e}")
            return False
    
    def _connect_excel(self) -> bool:
        """Подключиться к Excel файлу"""
        try:
            # Пробуем openpyxl для .xlsx
            if self.filepath.endswith('.xlsx'):
                import openpyxl
                self._workbook = openpyxl.load_workbook(self.filepath, read_only=True, data_only=True)
                self._sheet = self._workbook.active
            else:
                # Для .xls нужен xlrd
                import xlrd
                self._workbook = xlrd.open_workbook(self.filepath)
                self._sheet = self._workbook.sheet_by_index(0)
            
            # Читаем заголовки (первая строка)
            self._read_headers()
            
            self.is_connected = True
            return True
            
        except ImportError as e:
            print(f"Отсутствует библиотека: {e}")
            print("Установите: pip install openpyxl xlrd")
            return False
        except Exception as e:
            print(f"Ошибка чтения Excel: {e}")
            return False
    
    def _connect_csv(self) -> bool:
        """Подключиться к CSV файлу"""
        try:
            import csv
            
            with open(self.filepath, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                self.columns = reader.fieldnames or []
                self.data = list(reader)
            
            self.is_connected = True
            return True
            
        except Exception as e:
            print(f"Ошибка чтения CSV: {e}")
            return False
    
    def _read_headers(self):
        """Прочитать заголовки колонок"""
        if self._sheet is None:
            return
        
        try:
            if hasattr(self._sheet, 'iter_rows'):
                # openpyxl
                for row in self._sheet.iter_rows(min_row=1, max_row=1, values_only=True):
                    self.columns = [str(cell) if cell is not None else f"Column_{i}" 
                                   for i, cell in enumerate(row)]
                    break
            else:
                # xlrd
                row = self._sheet.row_values(0)
                self.columns = [str(cell) if cell else f"Column_{i}" 
                               for i, cell in enumerate(row)]
        except Exception as e:
            print(f"Ошибка чтения заголовков: {e}")
            self.columns = []
    
    def _read_all_data(self):
        """Прочитать все данные"""
        if self.data:
            return  # Уже прочитано для CSV
        
        try:
            if hasattr(self._sheet, 'iter_rows'):
                # openpyxl
                for row_idx, row in enumerate(self._sheet.iter_rows(
                    min_row=2,  # Пропускаем заголовок
                    values_only=True
                ), start=2):
                    if all(cell is None for cell in row):
                        continue
                    
                    record = {}
                    for col_idx, cell in enumerate(row):
                        if col_idx < len(self.columns):
                            col_name = self.columns[col_idx]
                            record[col_name] = cell if cell is not None else ""
                    
                    record['_row_index'] = row_idx
                    self.data.append(record)
            else:
                # xlrd
                for row_idx in range(1, self._sheet.nrows):
                    row = self._sheet.row_values(row_idx)
                    if all(not cell for cell in row):
                        continue
                    
                    record = {}
                    for col_idx, cell in enumerate(row):
                        if col_idx < len(self.columns):
                            col_name = self.columns[col_idx]
                            record[col_name] = cell if cell else ""
                    
                    record['_row_index'] = row_idx + 1  # 1-based
                    self.data.append(record)
                    
        except Exception as e:
            print(f"Ошибка чтения данных: {e}")
    
    def find_row(self, column: str, value: str) -> Optional[DatabaseRecord]:
        """Найти первую строку по значению"""
        self._read_all_data()
        
        for record_dict in self.data:
            cell_value = record_dict.get(column, "")
            if str(cell_value).strip() == str(value).strip():
                return DatabaseRecord(
                    row_index=record_dict.get('_row_index', 0),
                    data={k: v for k, v in record_dict.items() if k != '_row_index'}
                )
        
        return None
    
    def find_all_rows(self, column: str, value: str) -> List[DatabaseRecord]:
        """Найти все строки по значению"""
        self._read_all_data()
        
        results = []
        for record_dict in self.data:
            cell_value = record_dict.get(column, "")
            if str(cell_value).strip() == str(value).strip():
                results.append(DatabaseRecord(
                    row_index=record_dict.get('_row_index', 0),
                    data={k: v for k, v in record_dict.items() if k != '_row_index'}
                ))
        
        return results
    
    def get_column_values(
        self, 
        column: str, 
        filter_column: str = None,
        filter_value: str = None
    ) -> List[str]:
        """Получить все значения из колонки"""
        self._read_all_data()
        
        values = []
        for record_dict in self.data:
            # Применяем фильтр если указан
            if filter_column and filter_value:
                if str(record_dict.get(filter_column, "")).strip() != str(filter_value).strip():
                    continue
            
            cell_value = record_dict.get(column, "")
            if cell_value is not None and str(cell_value).strip():
                values.append(str(cell_value).strip())
        
        return list(dict.fromkeys(values))  # Удалить дубликаты с сохранением порядка
    
    def get_row_by_index(self, row_index: int) -> Optional[DatabaseRecord]:
        """Получить строку по индексу"""
        self._read_all_data()
        
        for record_dict in self.data:
            if record_dict.get('_row_index') == row_index:
                return DatabaseRecord(
                    row_index=row_index,
                    data={k: v for k, v in record_dict.items() if k != '_row_index'}
                )
        
        return None
    
    def close(self):
        """Закрыть подключение"""
        if self._workbook:
            if hasattr(self._workbook, 'close'):
                self._workbook.close()
            self._workbook = None
        
        if self._excel_app:
            try:
                self._excel_app.Quit()
            except:
                pass
            self._excel_app = None
        
        self.is_connected = False
        self.data.clear()
        self.columns.clear()
    
    def get_columns(self) -> List[str]:
        """Получить список колонок"""
        if not self.is_connected:
            self.connect()
        return self.columns
    
    def get_row_count(self) -> int:
        """Получить количество строк"""
        self._read_all_data()
        return len(self.data)
