"""
Реализация сервисов для macOS
"""

import logging
from backend.core import IMouseService, IKeyboardService, IScreenService, IDatabaseService, Coordinates, Color
from typing import Optional, Dict, Any, List
import os
from pathlib import Path

logger = logging.getLogger(__name__)


class MouseService(IMouseService):
    """Сервис мыши на основе pynput"""

    def __init__(self):
        self._mouse = None
        self._init_pynput()

    def _init_pynput(self):
        """Инициализация pynput"""
        try:
            from pynput.mouse import Controller
            self._mouse = Controller()
        except Exception as e:
            logger.exception("Ошибка инициализации pynput mouse")
            self._mouse = None

    def get_position(self) -> Coordinates:
        """Получить позицию мыши"""
        if self._mouse:
            try:
                x, y = self._mouse.position
                return Coordinates(int(x), int(y))
            except Exception:
                pass
        return Coordinates(0, 0)

    def move_to(self, x: int, y: int, duration_ms: int = 0) -> None:
        """Переместить мышь"""
        if self._mouse:
            try:
                if duration_ms > 0:
                    # Плавное перемещение
                    import time
                    start_x, start_y = self._mouse.position
                    steps = max(int(duration_ms / 16), 1)
                    for i in range(steps):
                        t = i / steps
                        new_x = start_x + (x - start_x) * t
                        new_y = start_y + (y - start_y) * t
                        self._mouse.position = (new_x, new_y)
                        time.sleep(0.016)
                
                self._mouse.position = (x, y)
                return
            except Exception as e:
                logger.exception("Ошибка перемещения мыши")

    def click(self, button: str = "left") -> None:
        """Клик мышью"""
        if self._mouse:
            try:
                from pynput.mouse import Button
                btn = Button.left
                if button == "right":
                    btn = Button.right
                elif button == "middle":
                    btn = Button.middle
                self._mouse.click(btn, 1)
                return
            except Exception as e:
                logger.exception("Ошибка клика мыши")


class KeyboardService(IKeyboardService):
    """Сервис клавиатуры на основе pynput"""

    def __init__(self):
        self._keyboard = None
        self._init_pynput()

    def _init_pynput(self):
        """Инициализация pynput"""
        try:
            from pynput.keyboard import Controller
            self._keyboard = Controller()
        except Exception as e:
            logger.exception("Ошибка инициализации pynput keyboard")
            self._keyboard = None

    def press(self, key: str) -> None:
        """Нажать клавишу"""
        if self._keyboard:
            try:
                if "+" in key:
                    self.press_hotkey([k.strip() for k in key.split("+") if k.strip()])
                    return
                key_obj = self._parse_key(key)
                self._keyboard.press(key_obj)
                self._keyboard.release(key_obj)
                return
            except Exception as e:
                logger.exception("Ошибка нажатия клавиши %s", key)

    def press_hotkey(self, keys: List[str]) -> None:
        """Нажать комбинацию клавиш"""
        if self._keyboard:
            try:
                from pynput.keyboard import Key

                modifiers = []
                main_key = None

                for k in keys:
                    key_obj = self._parse_key(k)
                    if key_obj in [Key.ctrl, Key.shift, Key.alt, Key.cmd]:
                        modifiers.append(key_obj)
                    else:
                        main_key = key_obj

                for mod in modifiers:
                    self._keyboard.press(mod)

                if main_key:
                    self._keyboard.press(main_key)
                    self._keyboard.release(main_key)

                for mod in reversed(modifiers):
                    self._keyboard.release(mod)

                return
            except Exception as e:
                logger.exception("Ошибка горячей клавиши %s", keys)

    def _parse_key(self, key: str):
        """Преобразовать строку в объект клавиши"""
        from pynput.keyboard import Key, KeyCode
        normalized_key = key.strip()

        special_keys = {
            'enter': Key.enter,
            'return': Key.enter,
            'esc': Key.esc,
            'escape': Key.esc,
            'tab': Key.tab,
            'backspace': Key.backspace,
            'delete': Key.delete,
            'home': Key.home,
            'end': Key.end,
            'pageup': Key.page_up,
            'pagedown': Key.page_down,
            'up': Key.up,
            'down': Key.down,
            'left': Key.left,
            'right': Key.right,
            'f1': Key.f1,
            'f2': Key.f2,
            'f3': Key.f3,
            'f4': Key.f4,
            'f5': Key.f5,
            'f6': Key.f6,
            'f7': Key.f7,
            'f8': Key.f8,
            'f9': Key.f9,
            'f10': Key.f10,
            'f11': Key.f11,
            'f12': Key.f12,
            'ctrl': Key.ctrl,
            'control': Key.ctrl,
            'shift': Key.shift,
            'alt': Key.alt,
            'cmd': Key.cmd,
            'win': Key.cmd,
            'space': Key.space,
        }

        key_lower = normalized_key.lower()
        if key_lower in special_keys:
            return special_keys[key_lower]

        if len(normalized_key) == 1:
            return KeyCode.from_char(normalized_key)

        if key_lower.startswith('f') and key_lower[1:].isdigit():
            f_num = int(key_lower[1:])
            if 1 <= f_num <= 24:
                return getattr(Key, f'f{f_num}')

        return KeyCode.from_char(normalized_key)


class ScreenService(IScreenService):
    """Сервис экрана на основе mss"""

    def __init__(self):
        self._sct = None
        self._monitor = None
        self._init_mss()

    def _init_mss(self):
        """Инициализация mss"""
        try:
            import mss
            self._sct = mss.mss()
            self._monitor = self._sct.monitors[0]
        except Exception as e:
            logger.exception("Ошибка инициализации mss")
            self._sct = None

    def get_pixel_color(self, x: int, y: int) -> Optional[Color]:
        """Получить цвет пикселя"""
        if self._sct:
            try:
                monitor = {"left": x, "top": y, "width": 1, "height": 1}
                screenshot = self._sct.grab(monitor)
                pixel = screenshot.pixel(0, 0)
                return Color(r=pixel[2], g=pixel[1], b=pixel[0])
            except Exception as e:
                logger.exception("Ошибка получения цвета пикселя")
        return None

    def take_screenshot(self, region: Optional[tuple] = None) -> Optional[str]:
        """Сделать скриншот"""
        if not self._sct:
            return None

        try:
            import time
            from PIL import Image

            if region:
                monitor = {
                    "left": region[0],
                    "top": region[1],
                    "width": region[2],
                    "height": region[3]
                }
            else:
                monitor = self._monitor

            screenshot = self._sct.grab(monitor)
            img = Image.frombytes("RGB", screenshot.size, screenshot.bgra, "raw", "BGRX")

            filename = Path.cwd() / f"screenshot_{int(time.time())}.png"
            img.save(filename)
            return str(filename)
        except Exception as e:
            logger.exception("Ошибка скриншота")
            return None

    def find_image(self, image_path: str, confidence: float = 0.9) -> Optional[Coordinates]:
        """Найти изображение"""
        try:
            if not self._sct or not self._monitor:
                return None
            import cv2
            import numpy as np
            from PIL import Image

            screenshot = self._sct.grab(self._monitor)
            screen_np = np.array(screenshot)
            screen_np = cv2.cvtColor(screen_np, cv2.COLOR_BGRA2BGR)

            template = cv2.imread(image_path)
            if template is None:
                return None

            result = cv2.matchTemplate(screen_np, template, cv2.TM_CCOEFF_NORMED)
            min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(result)

            if max_val >= confidence:
                h, w = template.shape[:2]
                center_x = int(max_loc[0] + w / 2)
                center_y = int(max_loc[1] + h / 2)
                return Coordinates(x=center_x, y=center_y)

            return None
        except Exception as e:
            logger.exception("Ошибка поиска изображения")
            return None

    def close(self) -> None:
        """Освободить ресурсы mss."""
        if self._sct is not None:
            try:
                self._sct.close()
            except Exception:
                logger.exception("Ошибка закрытия mss")
            finally:
                self._sct = None


class DatabaseService(IDatabaseService):
    """Сервис баз данных (Excel/CSV)"""

    def __init__(self):
        self.databases: Dict[str, Any] = {}

    def add_database(self, name: str, filepath: str) -> bool:
        """Добавить базу данных"""
        try:
            if not os.path.exists(filepath):
                return False

            ext = os.path.splitext(filepath)[1].lower()
            data = []
            columns = []

            if ext in ['.xlsx', '.xls']:
                data, columns = self._read_excel(filepath)
            elif ext == '.csv':
                data, columns = self._read_csv(filepath)
            else:
                return False

            self.databases[name] = {'data': data, 'columns': columns, 'filepath': filepath}
            return True
        except ImportError as e:
            logger.error("Не хватает зависимости для чтения БД: %s", e)
            return False
        except Exception as e:
            logger.exception("Ошибка добавления БД")
            return False

    def _read_excel(self, filepath: str) -> tuple:
        """Прочитать Excel"""
        data = []
        columns = []

        if filepath.endswith('.xlsx'):
            import openpyxl
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            try:
                ws = wb.active

                headers = [cell.value for cell in ws[1]]
                columns = [str(h) if h else f"Column_{i}" for i, h in enumerate(headers)]

                for row in ws.iter_rows(min_row=2, values_only=True):
                    if all(cell is None for cell in row):
                        continue
                    record = {columns[i]: cell for i, cell in enumerate(row) if i < len(columns)}
                    data.append(record)
            finally:
                wb.close()
        else:
            import xlrd
            wb = xlrd.open_workbook(filepath)
            ws = wb.sheet_by_index(0)

            columns = [str(cell) if cell else f"Column_{i}" for i, cell in enumerate(ws.row_values(0))]

            for row_idx in range(1, ws.nrows):
                row = ws.row_values(row_idx)
                if all(not cell for cell in row):
                    continue
                record = {columns[i]: cell for i, cell in enumerate(row) if i < len(columns)}
                data.append(record)

        return data, columns

    def _read_csv(self, filepath: str) -> tuple:
        """Прочитать CSV"""
        import csv
        data = []
        columns: List[str] = []
        encodings = ("utf-8", "utf-8-sig", "cp1251")

        for encoding in encodings:
            try:
                with open(filepath, 'r', encoding=encoding, newline='') as f:
                    reader = csv.DictReader(f)
                    columns = reader.fieldnames or []
                    data = list(reader)
                    return data, columns
            except UnicodeDecodeError:
                continue

        with open(filepath, 'r', encoding='utf-8', errors='replace', newline='') as f:
            reader = csv.DictReader(f)
            columns = reader.fieldnames or []
            data = list(reader)

        return data, columns

    def search(self, db_name: str, column: str, value: str) -> Optional[Dict[str, Any]]:
        """Поиск в БД"""
        db = self.databases.get(db_name)
        if not db:
            return None

        for record in db['data']:
            if str(record.get(column, '')).strip() == str(value).strip():
                return record

        return None

    def get_columns(self, db_name: str) -> List[str]:
        """Получить колонки"""
        db = self.databases.get(db_name)
        return db['columns'] if db else []
